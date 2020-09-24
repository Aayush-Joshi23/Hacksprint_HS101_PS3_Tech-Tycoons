# Importing all the necessary libraries
import numpy as np 
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl  
import time  
from openpyxl.styles import Alignment, Border, Side
import win32com.client
from pywintypes import com_error



# formatting data to process it
def format_dataframe(df):
    y = list(df.columns)
    x = list(df.columns)
    j = 1
    for i in range(5,len(x)-1):
        x[i] = "subject "+str(j)
        j += 1
    df.columns = x
    return df, y[5:10]


# Storing the names of all students in a txt file.
def names(df):
    file = open("names.txt", "w")
    student_names = list(df["Name"])
    for i in range(len(student_names)):
        file.write(student_names[i]+"\n")
    file.close()



# performing general analysis on the data.
# Finding the highest marks scored by a student in every subject
# Finding the average marks scored in every subject
def gen_analysis(df):
    s1 = max(df["subject 1"])
    s2 = max(df["subject 2"])
    s3 = max(df["subject 3"])
    s4 = max(df["subject 4"])
    s5 = max(df["subject 5"])
    max_marks = [s1, s2, s3, s4, s5]

    s1 = np.mean(df["subject 1"])
    s2 = np.mean(df["subject 2"])
    s3 = np.mean(df["subject 3"])
    s4 = np.mean(df["subject 4"])
    s5 =  np.mean(df["subject 5"])
    avg_marks = [s1, s2, s3, s4, s5]
    
    # Creating series for storing max and average marks
    s_1 = pd.Series(max_marks, index = ["subject 1", "subject 2", "subject 3", "subject 4", "subject 5"])
    s_2 = pd.Series(avg_marks, index = ["subject 1", "subject 2", "subject 3", "subject 4", "subject 5"])
    
    # Creating a data frame containing max and avg scores
    analysis = pd.DataFrame([s_1, s_2])
    # Changing the shape of the data frame 
    analysis = analysis.T
    analysis.rename(columns={0:"Max Marks", 1:"Average Marks"}, inplace = True)
    
    return analysis, avg_marks, max_marks



# Comparing marks of every student with max and avg marks and creating a comparison graph
# Joint Bar Graph
def student_perf_graph(student_marks, avg_marks, max_marks, subject_names):
    n_groups = 5
    data = [avg_marks, max_marks]
    data.append(student_marks)
    # create plot
    fig, ax = plt.subplots()
    index = np.arange(n_groups)
    bar_width = 0.25

    col_1 = plt.bar(index, data[0], bar_width, color='b', label='Average Marks')

    col_2 = plt.bar(index + bar_width, data[1], bar_width, color='g', label='Max Marks')

    col_3 = plt.bar(index + bar_width*2, data[2], bar_width, color='r', label='Student Marks')


    plt.xlabel('Subjects')
    plt.ylabel('Marks')
    plt.title('Comparison Graph')
    plt.xticks(index + bar_width, tuple(subject_names))
    plt.legend()
    
    # saving the plot as plot_1.png
    plt.savefig("plot_1.png", dpi = 150)
    plt.tight_layout()
    


# Creating the report card adding column with grades
# Adding an extra row with total marks
# Report currently contains no credentials
def report_card(student_marks, subject_names):
    subjects = ["subject 1", "subject 2", "subject 3", "subject 4", "subject 5"]
    min_marks = [4 for i in range(5)]
    max_marks = [10 for i in range(5)]

    student_marks.append(sum(student_marks)/5)
    grades = []
    for i in range(len(student_marks)):
        if student_marks[i] >= 9:
            grades.append("O")
        elif 9 > student_marks[i] >= 8:
            grades.append("A")
        elif 8 > student_marks[i] >= 7:
            grades.append("B")
        elif 7 > student_marks[i] >= 6:
            grades.append("C")
        elif 6 > student_marks[i] >= 5:
            grades.append("D")
        elif 5 > student_marks[i] >= 4:
            grades.append("P")
        else:
            grades.append("F")
        
    
    s1 = pd.Series(min_marks, index = subjects)
    s2 = pd.Series(max_marks, index = subjects)
    s3 = pd.Series(student_marks[:-1], index = subjects)
    
    report = pd.DataFrame((s1, s2, s3), index = ["Min", "Max", "Student"])
    report = report.T
    total = sum(list(report["Student"]))
    report = report.append({"Min" : sum(report["Min"]), "Max" : sum(report["Max"])
                           ,"Student" : total}, ignore_index = True)
    subject_names.append("Total")
    report["Subjects"] = subject_names
    report["Grade"] = grades
    report.set_index("Subjects", inplace = True)
    
    return report


# Function for adding detail to the report card
# Adding border to the data in the file.
def border_cell(sheet, cell_range):
    rows = sheet[cell_range]
    side = Side(border_style='thin', color="FF000000")

    rows = list(rows)  # we convert iterator to list for simplicity, but it's not memory efficient solution
    max_y = len(rows) - 1  # index of the last row
    for pos_y, cells in enumerate(rows):
        max_x = len(cells) - 1  # index of the last cell
        for pos_x, cell in enumerate(cells):
            border = Border(
                left=cell.border.left,
                right=cell.border.right,
                top=cell.border.top,
                bottom=cell.border.bottom
            )
            if pos_x == 0:
                border.left = side
            if pos_x == max_x:
                border.right = side
            if pos_y == 0:
                border.top = side
            if pos_y == max_y:
                border.bottom = side

            # set new border only if it's one of the edge cells
            if pos_x == 0 or pos_x == max_x or pos_y == 0 or pos_y == max_y:
                cell.border = border
    return sheet


# Formatting the final report card .
# Adding the credentials:
# Name/ Roll no/ College Code/ College Name/ Gender/ Attendance
# Inserting the comparison graph for each student.
# Returning the final percentage.
def final_format(r, x):
    mks = list(r["Student"])[-1]
    percentage = (mks / 50) * 100

    wb = openpyxl.load_workbook("report_card.xlsx")
    sheet = wb.active

    sheet = border_cell(sheet, "C1:G11")
    sheet = border_cell(sheet, "C4:G4")
    sheet = border_cell(sheet, "C1:D3")
    sheet = border_cell(sheet, "E5:E10")
    sheet = border_cell(sheet, "F5:F10")
    sheet = border_cell(sheet, "G5:G10")
    sheet = border_cell(sheet, "D5:D10")
    sheet = border_cell(sheet, "C10:G10")
    sheet = border_cell(sheet, "C1:G1")
    sheet = border_cell(sheet, "C2:G2")
    sheet = border_cell(sheet, "C3:G3")
    sheet = border_cell(sheet, "C11:D11")
    
    # Merging cells for adding credentials
    sheet.merge_cells('C1:D1')
    sheet.merge_cells('C11:D11')
    sheet.merge_cells('E1:G1')
    sheet.merge_cells('E2:G2')
    sheet.merge_cells('E3:G3')

    cell = sheet.cell(row=3, column=5)
    cell.value = "Name:   " + str(x[0])
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=1, column=5)
    cell.value = "College Name:   " + str(x[4])
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=2, column=5)
    cell.value = "Attendance:        " + str(x[-1]) + "%"
    cell.alignment = Alignment(horizontal='center', vertical='center')

    cell = sheet.cell(row=1, column=3)
    cell.value = "College Code:      " + str(x[3])
    cell.alignment = Alignment(horizontal='left', vertical='center')

    sheet['C2'] = "Gender"
    cell = sheet.cell(row=2, column=4)
    cell.value = x[2]
    cell.alignment = Alignment(horizontal='right', vertical='center')

    sheet['C3'] = "Roll No."
    sheet['D3'] = x[1]

    cell = sheet.cell(row=11, column=3)
    cell.value = "Percentage   ="
    cell.alignment = Alignment(horizontal='left', vertical='center')

    sheet["E11"] = str(percentage)+"%"

    for j in range(3, 8):
        for i in range(5, 11):
            cell = sheet.cell(row=i, column=j)
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # opening and inserting comparison graph of the student
    img = openpyxl.drawing.image.Image('plot_1.png')
    img.height = 250
    img.width = 475
    sheet.add_image(img, "B13")
    # saving the final report card as report_card.xlsx
    wb.save("report_card.xlsx")

    


# Converting the final report file i.e "report_card.xlsx" to pdf format
# saving the file with naming convention "student_rollno.pdf"
def convert_to_pdf(i):
    
    WB_PATH = r'C:\Users\HP\Desktop\Jupyter Notebooks\report_card.xlsx'
    
    
    PATH_TO_PDF = r'C:\Users\HP\Desktop\Jupyter Notebooks\student_{}.pdf'.format(i)


    excel = win32com.client.Dispatch("Excel.Application")

    excel.Visible = False

    try:
        print('Start conversion to PDF')

        
        wb = excel.Workbooks.Open(WB_PATH)

        
        ws_index_list = [1]
        wb.WorkSheets(ws_index_list).Select()

        
        wb.ActiveSheet.ExportAsFixedFormat(0, PATH_TO_PDF)
    except com_error as e:
        print('failed.')
    else:
        print('Succeeded.')
    finally:
        excel.Quit()
        


# Fianlly putting together all the functions and generating reports for every student .
def main():
    df = pd.read_excel("Required Format.xlsx")
    for i in range(len(df)):
        df = pd.read_excel("Required Format.xlsx")
        df, subject_names = format_dataframe(df)
        names(df)
        analysis, avg_marks, max_marks = gen_analysis(df)
        student = list(df.iloc[i])[5:10]
        graph = student_perf_graph(student, avg_marks, max_marks, subject_names)
        r = report_card(student, subject_names)
        r.to_excel("report_card.xlsx", sheet_name = "Sheet_1", startrow = 4)
        x = list(df.iloc[i])
        y = x[-1]
        x = x[:5]
        x.append(y)
        final_format(r, x)
        n = list(df.iloc[i])[1]
        convert_to_pdf(n)
main()






